import datetime
import sqlite3

import aiosqlite

from aiogram import F, Router, types
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import (Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, BufferedInputFile,
                           FSInputFile)
from aiogram.fsm.context import FSMContext
from openpyxl.styles import Alignment, Font

import states
import pandas as pd
import io
import db
from datetime import datetime
import logging

router = Router()
logger = logging.getLogger(__name__)
admin_kb= [
        [InlineKeyboardButton(text="Редактировать данные доноров", callback_data='donor_edit')],
        [InlineKeyboardButton(text="Изменить информацию в боте", callback_data='bot_edit')],
        [InlineKeyboardButton(text="Просмотреть статистику", callback_data='view_statistics')],
        [InlineKeyboardButton(text="Ответить на вопросы", callback_data='reply_to_questions')],
        [InlineKeyboardButton(text="Сделать рассылку", callback_data='newsletter'),
         InlineKeyboardButton(text="Создать мероприятие", callback_data='create_event')],
        [InlineKeyboardButton(text="Загрузить статистику", callback_data='upload_statistics'),
         InlineKeyboardButton(text="Получить статистику", callback_data='get_statistics')],
    ]

keyboard = InlineKeyboardMarkup(inline_keyboard=admin_kb)

"""
Два хендлера админ-панели со всеми кнопками
"""
@router.message(Command('admin'), states.IsAdmin())
async def admin_command(message: Message) -> None:
    await message.answer('Добро пожаловать в панель организатора!\n'
                         'Выберите желаемое действие:', reply_markup=keyboard)

@router.callback_query(F.data == "admin_menu")
async def back_to_admin(callback: CallbackQuery):
    await callback.message.edit_text('Добро пожаловать в панель организатора!\n'
                         'Выберите желаемое действие:', reply_markup=keyboard)




"""
Меню выбора редактирования пользователя добавить/изменить
"""
kb_edit_return = [
    [InlineKeyboardButton(text="Редактировать по телефону", callback_data='edit_by_phone')],
    [InlineKeyboardButton(text="Редактировать по ФИО", callback_data='edit_by_name')],
    [InlineKeyboardButton(text="Удалить по телефону", callback_data='delete_by_phone')],
    [InlineKeyboardButton(text="Добавить пользователя", callback_data='add_user')],
    [InlineKeyboardButton(text="Импорт из Excel", callback_data='import_from_excel')],
    [InlineKeyboardButton(text="Вернуться", callback_data='admin_menu')]
]
keyboard_edit_return = InlineKeyboardMarkup(inline_keyboard=kb_edit_return)


# Состояния для FSM
class EditStates(StatesGroup):
    waiting_phone = State()
    waiting_name = State()
    waiting_delete_phone = State()
    waiting_excel = State()
    editing_field = State()


"""
Меню выбора редактирования пользователя
"""


@router.callback_query(F.data == "donor_edit")
async def donor_edit(callback: CallbackQuery):
    await callback.message.edit_text(
        "Выберите действие с пользователями:",
        reply_markup=keyboard_edit_return
    )


"""
Редактирование по телефону
"""


@router.callback_query(F.data == "edit_by_phone")
async def edit_by_phone_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введите номер телефона пользователя для редактирования:")
    await state.set_state(EditStates.waiting_phone)
    await callback.answer()


@router.message(EditStates.waiting_phone)
async def edit_by_phone_process(message: Message, state: FSMContext):
    phone = message.text
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            cursor = await con.execute(
                "SELECT * FROM Donors WHERE Phone = ?",
                (phone,)
            )
            donor = await cursor.fetchone()

            if donor:
                await state.update_data(donor_id=donor[0])
                await show_donor_info(message, donor)
                await message.answer(
                    "Какое поле хотите изменить?",
                    reply_markup=create_full_edit_keyboard(donor[0])
                )
            else:
                await message.answer("Пользователь с таким телефоном не найден.")

    except Exception as e:
        await message.answer(f"Ошибка: {str(e)}")
    finally:
        await state.set_state(EditStates.editing_field)


"""
Редактирование по ФИО
"""


@router.callback_query(F.data == "edit_by_name")
async def edit_by_name_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введите ФИО пользователя для поиска:")
    await state.set_state(EditStates.waiting_name)
    await callback.answer()


@router.message(EditStates.waiting_name)
async def edit_by_name_process(message: Message, state: FSMContext):
    name = message.text
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            cursor = await con.execute(
                "SELECT * FROM Donors WHERE Name LIKE ?",
                (f"%{name}%",)
            )
            donors = await cursor.fetchall()

            if donors:
                if len(donors) > 1:
                    kb = []
                    for donor in donors:
                        kb.append([InlineKeyboardButton(
                            text=f"{donor[1]} ({donor[8]})",
                            callback_data=f"select_donor_{donor[0]}"
                        )])

                    await message.answer(
                        "Найдено несколько пользователей. Выберите нужного:",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
                    )
                else:
                    donor = donors[0]
                    await state.update_data(donor_id=donor[0])
                    await show_donor_info(message, donor)
                    await message.answer(
                        "Какое поле хотите изменить?",
                        reply_markup=create_full_edit_keyboard(donor[0])
                    )
                    await state.set_state(EditStates.editing_field)
            else:
                await message.answer("Пользователи с таким ФИО не найдены.")

    except Exception as e:
        await message.answer(f"Ошибка: {str(e)}")


"""
Обработчик выбора пользователя из списка
"""


@router.callback_query(F.data.startswith("select_donor_"))
async def select_donor(callback: CallbackQuery, state: FSMContext):
    donor_id = int(callback.data.split("_")[2])
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            cursor = await con.execute(
                "SELECT * FROM Donors WHERE donorID = ?",
                (donor_id,)
            )
            donor = await cursor.fetchone()

            if donor:
                await state.update_data(donor_id=donor[0])
                await show_donor_info(callback.message, donor)
                await callback.message.answer(
                    "Какое поле хотите изменить?",
                    reply_markup=create_full_edit_keyboard(donor[0])
                )
                await state.set_state(EditStates.editing_field)
            else:
                await callback.message.answer("Пользователь не найден.")

    except Exception as e:
        await callback.message.answer(f"Ошибка: {str(e)}")
    finally:
        await callback.answer()


"""
Функция для отображения информации о доноре
"""


async def show_donor_info(message: Message, donor):
    await message.answer(
        f"🔹 <b>Информация о пользователе:</b>\n\n"
        f"<b>ID:</b> {donor[0]}\n"
        f"<b>ФИО:</b> {donor[1]}\n"
        f"<b>Группа:</b> {donor[2]}\n"
        f"<b>Гаврилова:</b> {'Да' if donor[3] else 'Нет'}\n"
        f"<b>ФМБА:</b> {'Да' if donor[4] else 'Нет'}\n"
        f"<b>Последняя Гаврилова:</b> {donor[5]}\n"
        f"<b>Последняя ФМБА:</b> {donor[6]}\n"
        f"<b>Контакты:</b> {donor[7]}\n"
        f"<b>Телефон:</b> {donor[8]}\n"
        f"<b>Чужой:</b> {'Да' if donor[9] else 'Нет'}"
    )


"""
Функция для создания полной клавиатуры редактирования
"""


def create_full_edit_keyboard(donor_id):
    buttons = [
        [InlineKeyboardButton(text="ФИО", callback_data=f"edit_field_name_{donor_id}")],
        [InlineKeyboardButton(text="Группа", callback_data=f"edit_field_group_{donor_id}")],
        [InlineKeyboardButton(text="Гаврилова", callback_data=f"edit_field_gavrilova_{donor_id}")],
        [InlineKeyboardButton(text="ФМБА", callback_data=f"edit_field_fmba_{donor_id}")],
        [InlineKeyboardButton(text="Последняя Гаврилова", callback_data=f"edit_field_lastgav_{donor_id}")],
        [InlineKeyboardButton(text="Последняя ФМБА", callback_data=f"edit_field_lastfmba_{donor_id}")],
        [InlineKeyboardButton(text="Контакты", callback_data=f"edit_field_contacts_{donor_id}")],
        [InlineKeyboardButton(text="Телефон", callback_data=f"edit_field_phone_{donor_id}")],
        [InlineKeyboardButton(text="Чужой", callback_data=f"edit_field_stranger_{donor_id}")],
        [InlineKeyboardButton(text="Отмена", callback_data="donor_edit")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


"""
Обработчики редактирования полей
"""


@router.callback_query(F.data.startswith("edit_field_"))
async def edit_field_start(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    field_name = parts[2]
    donor_id = int(parts[3])

    await state.update_data(editing_field=field_name, donor_id=donor_id)

    field_descriptions = {
        "name": "Введите новое ФИО:",
        "group": "Введите новую группу:",
        "gavrilova": "Изменить статус Гаврилова (1 - да, 0 - нет):",
        "fmba": "Изменить статус ФМБА (1 - да, 0 - нет):",
        "lastgav": "Введите дату последней Гаврилова:",
        "lastfmba": "Введите дату последней ФМБА:",
        "contacts": "Введите новые контакты:",
        "phone": "Введите новый телефон:",
        "stranger": "Изменить статус Чужой (1 - да, 0 - нет):"
    }

    await callback.message.answer(field_descriptions[field_name])
    await state.set_state(EditStates.editing_field)
    await callback.answer()


@router.message(EditStates.editing_field)
async def edit_field_process(message: Message, state: FSMContext):
    data = await state.get_data()
    field_name = data['editing_field']
    donor_id = data['donor_id']
    new_value = message.text

    try:
        # Для числовых полей преобразуем значение
        if field_name in ['gavrilova', 'fmba', 'stranger']:
            new_value = 1 if new_value.lower() in ['1', 'да', 'yes', 'true'] else 0

        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            # Динамическое формирование SQL запроса
            await con.execute(
                f"UPDATE Donors SET {field_name.capitalize()} = ? WHERE donorID = ?",
                (new_value, donor_id)
            )
            await con.commit()

            # Получаем обновленные данные
            cursor = await con.execute(
                "SELECT * FROM Donors WHERE donorID = ?",
                (donor_id,)
            )
            donor = await cursor.fetchone()

            await message.answer("✅ Поле успешно обновлено!")
            await show_donor_info(message, donor)
            await message.answer(
                "Хотите изменить ещё что-то?",
                reply_markup=create_full_edit_keyboard(donor_id)
            )

    except Exception as e:
        await message.answer(f"❌ Ошибка при обновлении: {str(e)}")

    await state.update_data(editing_field=None)


"""
Удаление по телефону
"""


@router.callback_query(F.data == "delete_by_phone")
async def delete_by_phone_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введите номер телефона пользователя для удаления:")
    await state.set_state(EditStates.waiting_delete_phone)
    await callback.answer()


@router.message(EditStates.waiting_delete_phone)
async def delete_by_phone_process(message: Message, state: FSMContext):
    phone = message.text
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            cursor = await con.execute(
                "SELECT Name FROM Donors WHERE Phone = ?",
                (phone,)
            )
            donor = await cursor.fetchone()

            if donor:
                await con.execute(
                    "DELETE FROM Donors WHERE Phone = ?",
                    (phone,)
                )
                await con.commit()
                await message.answer(f"Пользователь {donor[0]} удалён.")
            else:
                await message.answer("Пользователь с таким телефоном не найден.")

    except Exception as e:
        await message.answer(f"Ошибка при удалении: {str(e)}")
    finally:
        await state.clear()


"""
Импорт из Excel
"""


@router.callback_query(F.data == "import_from_excel")
async def import_from_excel_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "Пожалуйста, отправьте Excel файл с данными пользователей.\n"
        "Формат файла должен соответствовать полям таблицы Donors."
    )
    await state.set_state(EditStates.waiting_excel)
    await callback.answer()


@router.message(EditStates.waiting_excel, F.document)
async def import_from_excel_process(message: Message, state: FSMContext):
    try:
        file = await message.bot.get_file(message.document.file_id)
        file_path = await message.bot.download_file(file.file_path)

        # Читаем Excel с указанием типов данных
        df = pd.read_excel(file_path, dtype={
            'Phone': str,
            'LastGavrilov': str,
            'LastFMBA': str,
            'GroupID': str
        })

        # Обработка телефонов
        df['Name'] = df['Name'].str.strip()
        df['Phone'] = df['Phone'].apply(clean_phone_number)

        # Обработка GroupID - первая буква заглавная, остальные строчные
        df['GroupID'] = df['GroupID'].str.strip().str.capitalize()

        # Обработка дат
        date_columns = ['LastGavrilov', 'LastFMBA']
        for col in date_columns:
            if col in df.columns:
                df[col] = df[col].apply(format_date)

        required_columns = ['Name', 'Phone']
        if not all(col in df.columns for col in required_columns):
            raise ValueError("Файл должен содержать колонки 'Name' и 'Phone'")

        async with aiosqlite.connect(db.DATABASE_NAME) as con:
            for _, row in df.iterrows():
                await con.execute(
                    """INSERT OR REPLACE INTO Donors 
                    (Name, GroupID, Gavrilova, FMBA, LastGavrilov, LastFMBA, Contacts, Phone, Stranger) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        row.get('Name'),
                        row.get('GroupID', ''),
                        int(row.get('Gavrilova', 0)),
                        int(row.get('FMBA', 0)),
                        row.get('LastGavrilov', ''),
                        row.get('LastFMBA', ''),
                        row.get('Contacts', ''),
                        row['Phone'],
                        int(row.get('Stranger', 0))
                    )
                )
            await con.commit()

        await message.answer(f"✅ Успешно импортировано {len(df)} пользователей.\n")

    except Exception as e:
        await message.answer(f"❌ Ошибка при импорте: {str(e)}")
    finally:
        await state.clear()


# Вспомогательные функции
def clean_phone_number(phone):
    """Приводит номер телефона к формату +79858920529"""
    if pd.isna(phone):
        return ''

    phone = str(phone)
    # Удаляем все нецифровые символы
    digits = ''.join(filter(str.isdigit, phone))

    # Добавляем +7 для российских номеров
    if digits.startswith('7') and len(digits) == 11:
        return f"+{digits}"
    elif digits.startswith('8') and len(digits) == 11:
        return f"+7{digits[1:]}"
    else:
        return f"+{digits}" if digits else ''


def format_date(date_str):
    """Приводит дату к формату dd-mm-yyyy"""
    if pd.isna(date_str) or not str(date_str).strip():
        return ''

    date_str = str(date_str).strip()

    try:
        # Пробуем разные форматы дат
        if '-' in date_str and len(date_str.split('-')[0]) == 2:
            # Уже в формате dd-mm-yyyy
            return date_str

        # Парсим дату из других форматов
        date_formats = [
            '%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y',
            '%d%m%Y', '%Y%m%d', '%d-%b-%y'
        ]

        for fmt in date_formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d-%m-%Y')
            except ValueError:
                continue

        return date_str  # Если не удалось распарсить, оставляем как есть
    except:
        return date_str  # В случае ошибки оставляем оригинальное значение


"""
Добавление нового пользователя по одному или списком. Добавление в таблицу Donors
"""
@router.callback_query(F.data == 'add_user')
async def start_add_donors(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "Введите данные доноров (каждая запись с новой строки):\n"
        "Формат: Фамилия,Имя,Отчество,Сотрудник/Группа студента,Телефон\n"
        "Необязательные поля (через запятую):\n"
        "баллы Гаврилова,баллы FMBA,Последняя дотация в центре Гаврилова,\n"
        "Последняя дотация в центре FMBA,Контакты,Внешний донор(0 - нет/1 - да)\n"
        "Тг айди\n\n"
        "Пример минимальных данных:\n"
        "Иванов,Иван,Иванович,Сотрудник,+79161234567\n\n"
        "Пример полных данных:\n"
        "Петров,Пётр,Петрович,Сотрудник,+79261234567,0,1,,2023-02-20,тел.123-456,0,1194604421"
    )
    await state.set_state(states.AddDonorsState.waiting_for_text)

@router.message(states.AddDonorsState.waiting_for_text)
async def handle_text_list(message: Message, state: FSMContext):
    lines = message.text.split('\n')
    added = 0

    async with aiosqlite.connect(db.DATABASE_NAME) as con, con.cursor() as cur:
        for i, line in enumerate(lines, 1):
            try:
                if not line.strip():
                    continue

                parts = [p.strip() for p in line.split(',')]
                print(parts)
                if len(parts) >= 3:
                    full_name = ' '.join(parts[:3])  # Объединяем фамилию, имя и отчество
                    group_id = parts[3]
                    phone = parts[4]

                    # Обработка необязательных полей
                    gavrilova = int(parts[5]) if len(parts) > 5 and parts[5] else 0
                    fmba = int(parts[6]) if len(parts) > 6 and parts[6] else 0
                    last_gavrilov = parts[7] if len(parts) > 7 else None
                    last_fmba = parts[8] if len(parts) > 8 else None
                    contacts = parts[9] if len(parts) > 9 else None
                    stranger = int(parts[10]) if len(parts) > 10 and parts[10] else 0
                    donorID = int(parts[11]) if len(parts) > 11 and parts[11] else 0

                    await con.execute(
                        """INSERT OR REPLACE INTO Donors
                        (Name, GroupID, Phone,
                         Gavrilova, FMBA, LastGavrilov, LastFMBA, Contacts, Stranger, donorID)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (full_name, group_id, phone,
                         gavrilova, fmba, last_gavrilov, last_fmba, contacts, stranger, donorID)
                    )
                    added += 1
            except Exception:
                continue

        await con.commit()

    await message.answer(f"✅ Добавлено {added} доноров")
    await state.clear()


kb_return = [
    [InlineKeyboardButton(text="Вернуться", callback_data='admin_menu')]
]
keyboard_return = InlineKeyboardMarkup(inline_keyboard=kb_return)


"""
Изменения информации о боте (заглушка)
"""
@router.callback_query(F.data == "bot_edit")
async def donor_edit(callback: CallbackQuery):
    await callback.message.edit_text("Выберите, что хотите изменить",
                                     reply_markup=keyboard_return)

"""
Просмотр статистики просто зарегистрированных на дату и тех кто сдал кровь.
Также возвращает exel файл
"""
@router.callback_query(F.data == "view_statistics")
async def donor_edit(callback: CallbackQuery):
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            # Запрос с преобразованием форматов дат
            query = """
                SELECT 
                    dd.Date AS dd_date,
                    dd.donor_center AS center,
                    COUNT(dd_data.donorID) AS total_registrations,
                    SUM(CASE WHEN dd_data.complete = 1 THEN 1 ELSE 0 END) AS completed_donations
                FROM DD dd
                LEFT JOIN donors_data dd_data ON 
                    -- Преобразуем дату из DD в формат YYYY-MM-DD для сравнения
                    CASE 
                        WHEN length(dd.Date) = 8 THEN 
                            substr(dd.Date, 5, 4) || '-' || substr(dd.Date, 3, 2) || '-' || substr(dd.Date, 1, 2)
                        WHEN length(dd.Date) = 10 AND dd.Date LIKE '__-__-____' THEN 
                            substr(dd.Date, 7, 4) || '-' || substr(dd.Date, 4, 2) || '-' || substr(dd.Date, 1, 2)
                        ELSE dd.Date
                    END = dd_data.Date
                GROUP BY dd.Date, dd.donor_center
                ORDER BY 
                    CASE 
                        WHEN length(dd.Date) = 8 THEN 
                            substr(dd.Date, 5, 4) || substr(dd.Date, 3, 2) || substr(dd.Date, 1, 2)
                        WHEN length(dd.Date) = 10 AND dd.Date LIKE '__-__-____' THEN 
                            substr(dd.Date, 7, 4) || substr(dd.Date, 4, 2) || substr(dd.Date, 1, 2)
                        ELSE dd.Date
                    END DESC,
                    dd.donor_center
                """

            cursor = await conn.execute(query)
            stats = await cursor.fetchall()

            if not stats:
                await callback.message.answer("Нет данных о донорских акциях")
                return

            # Формируем текстовый отчет
            report = "📊 Статистика по донорским акциям:\n\n"
            report += "Дата       | Центр      | Регистраций | Завершено\n"
            report += "-----------------------------------------------\n"

            current_date = None
            for row in stats:
                # Форматируем дату из DD (исходный формат)
                raw_date = row[0]
                try:
                    if len(raw_date) == 8 and raw_date.isdigit():  # Формат ДДММГГГГ
                        formatted_date = f"{raw_date[:2]}.{raw_date[2:4]}.{raw_date[4:]}"
                    elif len(raw_date) == 10 and '-' in raw_date:  # Формат ДД-ММ-ГГГГ
                        formatted_date = raw_date.replace('-', '.')
                    else:
                        formatted_date = raw_date
                except:
                    formatted_date = raw_date

                # Добавляем пустую строку между разными датами
                if current_date != row[0]:
                    current_date = row[0]
                    report += "\n"

                report += (f"{formatted_date.ljust(10)} | {str(row[1]).ljust(10)} | "
                           f"{str(row[2]).center(11)} | {row[3]}\n")

            # Создаем Excel-отчет
            excel_buffer = io.BytesIO()

            # Преобразуем даты для Excel
            excel_data = []
            for row in stats:
                raw_date = row[0]
                try:
                    if len(raw_date) == 8 and raw_date.isdigit():  # Формат ДДММГГГГ
                        excel_date = f"{raw_date[:2]}.{raw_date[2:4]}.{raw_date[4:]}"
                    elif len(raw_date) == 10 and '-' in raw_date:  # Формат ДД-ММ-ГГГГ
                        excel_date = raw_date.replace('-', '.')
                    else:
                        excel_date = raw_date
                except:
                    excel_date = raw_date

                excel_data.append([excel_date, row[1], row[2], row[3]])

            df = pd.DataFrame(excel_data, columns=['Date', 'Центр', 'Регистрации', 'Завершено'])
            df.to_excel(excel_buffer, index=False, engine='openpyxl')
            excel_buffer.seek(0)

            # Отправляем оба варианта
            await callback.message.answer(report)
            await callback.message.answer_document(
                BufferedInputFile(
                    excel_buffer.getvalue(),
                    filename="Статистика_акций.xlsx"
                ),
                caption="Подробная статистика (регистрации/завершено)"
            )

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при получении статистики: {str(e)}")
    finally:
        await callback.answer()


"""
Ответы на вопросы пользователей (заглушка)
"""
@router.callback_query(F.data == "reply_to_questions")
async def donor_edit(callback: CallbackQuery):
    await callback.message.edit_text("Вопрос:",
                                     reply_markup=keyboard_return)

"""
Рассылка пользователям, зарегестрированным на ближайшую донацию
"""
@router.callback_query(F.data == "newsletter")
async def newsletter_menu(callback: types.CallbackQuery):
    buttons = [
        [types.InlineKeyboardButton(
            text="Рассылка зарегистрированным на ближайший ДД",
            callback_data="newsletter_nearest"
        )],
        [types.InlineKeyboardButton(
            text="Назад",
            callback_data="admin_menu"
        )]
    ]

    await callback.message.edit_text(
        text="<b>Меню рассылки</b>\nВыберите тип рассылки:",
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()

@router.callback_query(F.data == "newsletter_nearest")
async def newsletter_nearest_dd(callback: types.CallbackQuery, state: FSMContext):
    nearest_date = await get_nearest_future_date()
    if not nearest_date:
        await callback.message.edit_text(
            text="❌ Нет доступных дат для рассылки",
            reply_markup=types.InlineKeyboardMarkup(inline_keyboard=[
                [types.InlineKeyboardButton(text="Назад", callback_data="newsletter")]
            ]))
        return

    recipients = await get_recipients_for_date(nearest_date)
    if not recipients:
        await callback.message.edit_text(
            text=f"❌ На {nearest_date} нет зарегистрированных пользователей",
            reply_markup=types.InlineKeyboardMarkup(inline_keyboard=[
                [types.InlineKeyboardButton(text="Назад", callback_data="newsletter")]
            ]))
        return
    await state.update_data(
        newsletter_date=nearest_date,
        recipient_ids=recipients
    )
    await state.set_state(states.NewsletterStates.waiting_for_message)

    await callback.message.edit_text(
        text=f"📅 Ближайший ДД: <b>{nearest_date}</b>\n"
             f"👥 Зарегистрировано: <b>{len(recipients)}</b> человек\n\n"
             "Введите сообщение для рассылки или /cancel для отмены:",
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text="❌ Отмена", callback_data="newsletter")]
        ])
    )
    await callback.answer()

@router.message(states.NewsletterStates.waiting_for_message, F.text)
async def process_newsletter_message(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if not data.get('recipient_ids'):
        await message.answer("❌ Нет получателей для рассылки. Начните заново.")
        await state.clear()
        return

    try:
        sent_count = 0
        for user_id in data['recipient_ids']:
            try:
                await message.bot.send_message(
                    chat_id=user_id,
                    text=message.text
                )
                sent_count += 1
                await  datetime.time.sleep(0.1)
            except Exception as e:
                logger.error(f"Error sending to {user_id}: {e}")

        await message.answer(
            f"✅ Рассылка завершена!\n"
            f"Дата: <b>{data['newsletter_date']}</b>\n"
            f"Отправлено: <b>{sent_count}/{len(data['recipient_ids'])}</b>",
            parse_mode="HTML"
        )
    except Exception as e:
        logger.error(f"Newsletter error: {e}")
        await message.answer("❌ Ошибка при рассылке")
    finally:
        await state.clear()

@router.message(Command("cancel"))
async def cancel_newsletter(message: types.Message, state: FSMContext):
    """Отмена рассылки"""
    current_state = await state.get_state()
    if current_state is None:
        return

    await state.clear()
    await message.answer(
        "❌ Рассылка отменена",
        reply_markup=types.ReplyKeyboardRemove()
    )


"""
Создание нового мероприятия. Добавление в таблицу DD
"""
@router.callback_query(F.data == "create_event")
async def create_event(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Для создания мероприятия введите дату (dd-mm-yyyy) и "
                                     "центр крови через запятую.\n"
                                     "Каждое мероприятие на новой строчке\n"
                                     "Пример:\n"
                                     "24-03-2024,Гаврилова",
                                     reply_markup=keyboard_return)
    await state.set_state(states.EventState.waiting_for_event)

@router.message(states.EventState.waiting_for_event)
async def select_donor_field(message: Message, state: FSMContext):
    lines = message.text.split('\n')
    added = 0

    async with aiosqlite.connect(db.DATABASE_NAME) as con, con.cursor() as cur:
        for i, line in enumerate(lines, 1):
            try:
                if not line.strip():
                    continue

                parts = [p.strip() for p in line.split(',')]
                if len(parts) >= 2:
                    date = parts[0]
                    donor_center = parts[1]

                    await con.execute(
                        """INSERT OR REPLACE INTO DD
                        (Date, donor_center)
                        VALUES (?, ?)""",
                        (date, donor_center)
                    )
                    added += 1
            except Exception:
                continue

        await con.commit()

    await message.answer(f"✅ Добавлено {added} событий")
    await state.clear()



"""
Загрузка данных в таблицу Donors_date через exel файл
"""
@router.callback_query(F.data == "upload_statistics")
async def donor_edit(callback: CallbackQuery, state: FSMContext):
    # Путь к вашей картинке (замените на реальный путь)
    image_path = r"images/upload_stat.jpg"

    try:
        # Сначала отправляем картинку
        photo = FSInputFile(image_path)
        await callback.message.answer_photo(
            photo,
            caption="Пример файла для загрузки.\n"
                    "должны содержаться все столбцы"
        )

        # Затем отправляем текстовое сообщение с клавиатурой
        await callback.message.answer(
            "Для загрузки статистики отправьте файл в формате Excel.\n"
            "Принимаются только файлы формата .xlsx и .xls\n",
            reply_markup=keyboard_return
        )

        # Устанавливаем состояние
        await state.set_state(states.DocumentState.waiting_for_document)

    except Exception as e:
        await callback.message.answer(f"Ошибка загрузки инструкции: {str(e)}")
        await callback.message.answer(
            "Для загрузки статистики отправьте файл в формате Excel",
            reply_markup=keyboard_return
        )
        await state.set_state(states.DocumentState.waiting_for_document)

    finally:
        await callback.answer()


@router.message(states.DocumentState.waiting_for_document, F.document)
async def handle_excel_document(message: Message):
    # Проверяем, что это Excel-файл
    if not message.document.file_name.endswith(('.xlsx', '.xls', 'XLSX')):
        await message.answer("Пожалуйста, загрузите файл в формате Excel (.xlsx или .xls)")
        return

    try:
        file_bytes = await message.bot.download(message.document, destination=io.BytesIO())
        file_bytes.seek(0)

        try:
            df = pd.read_excel(file_bytes)
        except Exception as e:
            await message.answer(f"Ошибка чтения Excel-файла: {str(e)}")
            return

        required_columns = ['ФИО', 'Дата акции', 'ЦК', 'Статус', 'Тип', 'Завершено']
        if not all(col in df.columns for col in required_columns):
            missing = set(required_columns) - set(df.columns)
            await message.answer(f"В файле отсутствуют обязательные колонки: {', '.join(missing)}")
            return

        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            total_added = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    # Очищаем ФИО от лишних пробелов
                    name = str(row['ФИО']).strip()

                    # Преобразуем дату
                    date_str = pd.to_datetime(row['Дата акции']).strftime('%Y-%m-%d')

                    # Ищем донора в базе
                    cursor = await conn.execute(
                        "SELECT donorID FROM Donors WHERE Name = ?",
                        (name,)
                    )
                    donor = await cursor.fetchone()
                    if not donor:
                        errors.append(f"Донор {name} не найден в базе")
                        continue

                    donor_id = donor[0]

                    # Вставляем данные, включая имя
                    await conn.execute(
                        """INSERT OR REPLACE INTO donors_data 
                        (Date, donorID, Name, donor_status, donor_type, complete)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                        (
                            date_str,
                            donor_id,
                            name,  # Добавляем имя в запрос
                            int(row['Статус']),
                            int(row['Тип']),
                            int(row['Завершено'])
                        )
                    )
                    total_added += 1

                except Exception as e:
                    errors.append(f"Ошибка в строке {index + 1}: {str(e)}")
                    continue

            await conn.commit()

            # Формируем отчет
            report = f"✅ Успешно загружено: {total_added} записей"
            if errors:
                report += f"\n\nОшибки ({len(errors)}):\n" + "\n".join(errors[:5])  # Показываем первые 5 ошибок

            await message.answer(report)

    except Exception as e:
        await message.answer(f"❌ Ошибка обработки файла: {str(e)}")

kb_stat = [
    [InlineKeyboardButton(text='Список мероприятий', callback_data='export_dd' )],
    [InlineKeyboardButton(text='Список доноров', callback_data='export_donors' )],
    [InlineKeyboardButton(text='Списки зарегистрированных на событие', callback_data='export_donors_date' )],
    [InlineKeyboardButton(text='Вернуться', callback_data='admin_menu')],
]
keyboard_stat = InlineKeyboardMarkup(inline_keyboard=kb_stat)

"""
Хендлеры для получение файлов статистики из таблиц DD, Donors, donors_data
"""
@router.callback_query(F.data == "get_statistics")
async def donor_edit(callback: CallbackQuery):
    await callback.message.edit_text("Выберите желаемую информацию",
                                     reply_markup=keyboard_stat)


@router.callback_query(F.data == 'export_dd')
async def export_dd_table(callback: CallbackQuery):
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            # Правильное выполнение запроса с aiosqlite
            cursor = await conn.execute("SELECT * FROM DD")
            rows = await cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

            # Создаем DataFrame
            dd_df = pd.DataFrame(rows, columns=columns)

            # Создаем Excel файл в памяти
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                dd_df.to_excel(writer, index=False, sheet_name='Мероприятия')
            excel_buffer.seek(0)

            # Отправляем файл
            await callback.message.answer_document(
                BufferedInputFile(
                    excel_buffer.getvalue(),
                    filename="Мероприятия.xlsx"
                ),
                caption="Список мероприятий"
            )

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при экспорте мероприятий: {str(e)}")
    finally:
        await callback.answer()


@router.callback_query(F.data == 'export_donors')
async def export_donors_table(callback: CallbackQuery):
    try:
        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            cursor = await conn.execute("SELECT * FROM Donors")
            rows = await cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

            donors_df = pd.DataFrame(rows, columns=columns)

            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                donors_df.to_excel(writer, index=False, sheet_name='Доноры')
            excel_buffer.seek(0)

            await callback.message.answer_document(
                BufferedInputFile(
                    excel_buffer.getvalue(),
                    filename="Доноры.xlsx"
                ),
                caption="Список доноров"
            )

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при экспорте доноров: {str(e)}")
    finally:
        await callback.answer()


@router.callback_query(F.data == 'export_donors_date')
async def export_donors_by_date(callback: CallbackQuery):
    try:
        # Сначала получаем список доступных дат
        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            cursor = await conn.execute("SELECT DISTINCT Date FROM DD ORDER BY Date DESC")
            dates = await cursor.fetchall()

            if not dates:
                await callback.message.answer("Нет доступных мероприятий")
                return

            # Создаем клавиатуру с датами
            kb_dates = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=date[0], callback_data=f"export_date_{date[0]}")]
                for date in dates
            ])

            await callback.message.edit_text(
                "Выберите дату мероприятия:",
                reply_markup=kb_dates
            )

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при получении дат: {str(e)}")
    finally:
        await callback.answer()


@router.callback_query(F.data.startswith("export_date_"))
async def export_for_selected_date(callback: CallbackQuery):
    try:
        selected_date = callback.data.split("_")[2]

        async with aiosqlite.connect(db.DATABASE_NAME) as conn:
            # Получаем информацию о мероприятии
            event_cursor = await conn.execute(
                "SELECT donor_center FROM DD WHERE Date = ?",
                (selected_date,)
            )
            event_info = await event_cursor.fetchone()

            if not event_info:
                await callback.message.answer(f"Мероприятие на {selected_date} не найдено")
                return

            event_name = event_info[0]

            # Преобразуем формат даты для поиска в donors_data
            try:
                date_obj = datetime.strptime(selected_date, "%d-%m-%Y")
                db_date = date_obj.strftime("%Y-%m-%d")
            except ValueError:
                await callback.message.answer("Неверный формат даты. Используйте ДД-ММ-ГГГГ")
                return

            # Модифицированный запрос с включением статуса Complete
            query = """
            SELECT 
                d.*,
                dd.complete AS Статус
            FROM Donors d
            JOIN donors_data dd ON d.donorID = dd.donorID
            WHERE dd.Date = ?
            """
            cursor = await conn.execute(query, (db_date,))
            rows = await cursor.fetchall()

            if not rows:
                # Дополнительная проверка альтернативного формата даты
                alt_cursor = await conn.execute(query, (selected_date,))
                alt_rows = await alt_cursor.fetchall()

                if not alt_rows:
                    await callback.message.answer(
                        f"На мероприятии '{event_name}' ({selected_date}) нет зарегистрированных доноров"
                    )
                    return
                rows = alt_rows

            columns = [desc[0] for desc in cursor.description]

            # Создаем DataFrame и преобразуем статус в читаемый формат
            df = pd.DataFrame(rows, columns=columns)
            df['Статус'] = df['Статус'].apply(lambda x: 'Да' if x == 1 else 'Нет')

            # Создаем Excel файл с улучшенным форматированием
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Доноры')

                # Получаем объект листа для форматирования
                worksheet = writer.sheets['Доноры']

                # Добавляем информацию о мероприятии
                worksheet.cell(row=1, column=len(columns) + 1, value=f"Мероприятие: {event_name}")
                worksheet.cell(row=2, column=len(columns) + 1, value=f"Дата: {selected_date}")

                # Форматируем заголовки
                for col in worksheet.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                # Автонастройка ширины столбцов
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

            excel_buffer.seek(0)

            await callback.message.answer_document(
                BufferedInputFile(
                    excel_buffer.getvalue(),
                    filename=f"Список_доноров_{selected_date.replace('-', '_')}.xlsx"
                ),
                caption=f"Список доноров на {selected_date} ({event_name})\nСтатус: 'Да' - донация завершена, 'Нет' - не завершена"
            )

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при экспорте: {str(e)}")
    finally:
        await callback.answer()


async def get_nearest_future_date():
    try:
        conn = sqlite3.connect(db.DATABASE_NAME)
        cursor = conn.cursor()

        cursor.execute("SELECT date FROM DD")
        all_dates = [row[0] for row in cursor.fetchall()]

        future_dates = []
        current_date = datetime.now()

        for date_str in all_dates:
            try:
                date_obj = datetime.strptime(date_str, "%d-%m-%Y")
                if date_obj > current_date:
                    future_dates.append((date_obj, date_str))  # Сохраняем и объект, и строку
            except ValueError as e:
                logger.error(f"Ошибка парсинга даты {date_str}: {e}")

        if not future_dates:
            return None

        nearest_date_obj, nearest_date_str = min(future_dates, key=lambda x: x[0])
        return nearest_date_str

    except Exception as e:
        logger.error(f"Error getting nearest date: {e}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()


async def get_recipients_for_date(date_str):
    try:
        conn = sqlite3.connect(db.DATABASE_NAME)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT donorID FROM donors_data 
            WHERE Data = ?
        """, (date_str,))

        return [row[0] for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error getting recipients: {e}")
        return []
    finally:
        if 'conn' in locals():
            conn.close()


async def get_recipients_for_date(date):
    try:
        conn = sqlite3.connect(db.DATABASE_NAME)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT donorID FROM donors_data 
            WHERE Date = ?
        """, (date,))

        return [row[0] for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error getting recipients: {e}")
        return []
    finally:
        if 'conn' in locals():
            conn.close()
